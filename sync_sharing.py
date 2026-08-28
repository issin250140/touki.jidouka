#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sync_sharing.py
================
「共有管理.xlsx」の内容を読み取り、Renderの環境変数 SITE_USERS を自動で
同期する（Render REST APIを使用）。手動でRenderのダッシュボードを
操作しなくても、シートを保存するだけで反映されるようにするためのツール。

事前準備（最初の1回だけ）:
  1. https://dashboard.render.com/settings#api-keys を開く
     (右上のアカウントメニュー → Account Settings → API Keys でも可)
  2. 「Create API Key」で新しいAPIキーを作成し、表示された値をコピーする
     （このキーはこのスクリプトとClaudeどちらにも直接教えないでください。
      次の手順でご自身のパソコンだけに保存します）。
  3. PowerShellを開き、以下を実行する（<コピーした値> を実際の値に置き換える）:

     [Environment]::SetEnvironmentVariable("RENDER_API_KEY", "<コピーした値>", "User")

     実行後、一度PowerShellを閉じて開き直してください（環境変数の反映のため）。

使い方:
    python sync_sharing.py            # シートを見て、Render側と差分があれば同期し、再起動をかける
    python sync_sharing.py --check    # 何も変更せず、今シート上で共有中の人を表示するだけ

「共有を更新する.bat」をダブルクリックすれば、このスクリプトを起動できます。
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

SERVICE_ID = "srv-da85fdgn74is739j4ak0"
ENV_VAR_KEY = "SITE_USERS"
API_BASE = "https://api.render.com/v1"

SCRIPT_DIR = Path(__file__).resolve().parent
SHEET_PATH = SCRIPT_DIR / "共有管理.xlsx"
LOG_PATH = SCRIPT_DIR / "sync_log.txt"

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


def read_sheet_users() -> dict:
    """共有管理.xlsx を読み、「共有する」が「はい」の人だけを
    {ユーザー名: (名前, パスワード)} の形で返す。
    """
    wb = openpyxl.load_workbook(SHEET_PATH, data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "名前（ニックネーム可）":
            header_row = r
            break
    if header_row is None:
        raise RuntimeError(
            "シートのヘッダー行（1列目『名前（ニックネーム可）』）が見つかりませんでした。"
            "シートの列構成を変更していないか確認してください。"
        )

    result = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        username = ws.cell(row=r, column=2).value
        password = ws.cell(row=r, column=3).value
        share = ws.cell(row=r, column=4).value

        if isinstance(name, str) and name.startswith("（例）"):
            continue  # 説明用の例の行は無視
        if not username or not password:
            continue
        if str(share).strip() != "はい":
            continue

        result[str(username).strip()] = (
            str(name).strip() if name else "",
            str(password).strip(),
        )
    return result


def build_site_users_value(users: dict) -> str:
    return ",".join(f"{u}:{p}" for u, (_name, p) in sorted(users.items()))


def get_current_value(api_key: str):
    r = requests.get(
        f"{API_BASE}/services/{SERVICE_ID}/env-vars/{ENV_VAR_KEY}",
        headers={"Authorization": f"Bearer {api_key}"},
        timeout=15,
    )
    if r.status_code == 404:
        return None
    r.raise_for_status()
    return r.json().get("value")


def set_value(api_key: str, value: str) -> None:
    r = requests.put(
        f"{API_BASE}/services/{SERVICE_ID}/env-vars/{ENV_VAR_KEY}",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={"value": value},
        timeout=15,
    )
    r.raise_for_status()


def trigger_deploy(api_key: str) -> None:
    """環境変数の更新だけでは動作中のサーバーに反映されないため、
    コードの再ビルドをせず再起動だけする「deploy_only」でデプロイをかける。
    """
    r = requests.post(
        f"{API_BASE}/services/{SERVICE_ID}/deploys",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={"deployMode": "deploy_only"},
        timeout=15,
    )
    r.raise_for_status()


def log(message: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {message}"
    print(line)
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(line + "\n")
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="同期はせず、現在シート上で共有中の人を表示するだけ")
    args = parser.parse_args()

    if not SHEET_PATH.exists():
        log(f"エラー: {SHEET_PATH.name} が見つかりません（{SCRIPT_DIR} に置いてください）。")
        sys.exit(1)

    try:
        users = read_sheet_users()
    except PermissionError:
        log(
            f"エラー: {SHEET_PATH.name} を読み込めませんでした。"
            "Excelでこのファイルを開いたままになっていないか確認し、"
            "保存してから閉じて、もう一度お試しください。"
        )
        sys.exit(1)
    names = ", ".join(f"{name or username}" for username, (name, _p) in users.items()) or "(誰もいません)"

    if args.check:
        log(f"[確認のみ] 現在シート上で「共有する」になっている人: {names}")
        return

    api_key = os.environ.get("RENDER_API_KEY")
    if not api_key:
        log("エラー: 環境変数 RENDER_API_KEY が設定されていません。ファイル冒頭の説明を参照してください。")
        sys.exit(1)

    desired = build_site_users_value(users)

    try:
        current = get_current_value(api_key)
    except requests.exceptions.HTTPError as e:
        if e.response is not None and e.response.status_code == 401:
            log("エラー: RenderのAPIキーが正しくないようです。設定し直してください。")
        else:
            log(f"エラー: Renderとの通信に失敗しました（{e}）。")
        sys.exit(1)
    except requests.exceptions.RequestException as e:
        log(f"エラー: インターネットに接続できないか、Renderに繋がりませんでした（{e}）。")
        sys.exit(1)

    if current == desired:
        log(f"変更なし。現在共有中: {names}")
        return

    try:
        set_value(api_key, desired)
        trigger_deploy(api_key)
    except requests.exceptions.RequestException as e:
        log(f"エラー: 更新に失敗しました（{e}）。")
        sys.exit(1)

    log(f"更新し、反映のための再起動をかけました。現在共有中: {names}")
    log("（反映には30秒〜1分ほどかかります。すぐに開くと古い状態のことがあります）")


if __name__ == "__main__":
    main()
